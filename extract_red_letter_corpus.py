#!/usr/bin/env python3
"""
Extract the complete red-letter corpus (words of Jesus) from the public-domain
World English Bible (WEB) into the project's 12-column schema.

WHY THIS SCRIPT EXISTS
----------------------
The clean, machine-readable source that marks *which* words are Jesus's is the
WEB in USFX form (eBible.org), which tags his speech with <wj>...</wj>. That file
ships as a zip, so it must be downloaded on a networked machine. Everything else
(parsing, grouping, schema) is done here.

USAGE
-----
    pip install openpyxl
    # Option A: let the script download the WEB USFX zip itself
    python extract_red_letter_corpus.py --out jesus_full_red_letter.xlsx
    # Option B: point at a USFX .xml you already downloaded/extracted
    python extract_red_letter_corpus.py --usfx engwebpusfx.xml --out out.xlsx

The output matches the schema in jesus_sayings_dataset.xlsx. The four annotation
columns (Modern Rendering, Situational Context, Sentiment, Audience, Location,
Reason, Reasoning Move) are left blank for the tagging pass per the rubric.
"""
import argparse, io, sys, zipfile, urllib.request
import xml.etree.ElementTree as ET

USFX_ZIP_URL = "https://ebible.org/Scriptures/engwebp_usfx.zip"

# Book code -> (display name, traditional author). NT books that contain red letter.
BOOKS = {
    "MAT": ("Matthew", "Matthew (trad.)"),
    "MRK": ("Mark", "Mark (trad.)"),
    "LUK": ("Luke", "Luke (trad.)"),
    "JHN": ("John", "John (trad.)"),
    "ACT": ("Acts", "Luke (trad.)"),
    "REV": ("Revelation", "John (trad.)"),
}
BOOK_ORDER = ["MAT", "MRK", "LUK", "JHN", "ACT", "REV"]

def local(tag):
    return tag.split("}")[-1] if "}" in tag else tag

def get_usfx_bytes(args):
    if args.usfx:
        with open(args.usfx, "rb") as f:
            return f.read()
    print(f"Downloading {USFX_ZIP_URL} ...", file=sys.stderr)
    req = urllib.request.Request(USFX_ZIP_URL, headers={
        "User-Agent": "Mozilla/5.0 (compatible; red-letter-corpus/1.0)"
    })
    data = urllib.request.urlopen(req, timeout=120).read()
    zf = zipfile.ZipFile(io.BytesIO(data))
    xmlname = next((n for n in zf.namelist() if "usfx" in n.lower() and n.lower().endswith(".xml")),
                   next(n for n in zf.namelist() if n.lower().endswith(".xml")))
    print(f"Extracting {xmlname}", file=sys.stderr)
    return zf.read(xmlname)

def parse_wj(xml_bytes):
    """Return list of (book_code, chapter, verse, segment, text) for every <wj> span.
    `segment` increments at each block boundary (paragraph/poetry line) so that long
    discourses are split into their natural sense-units instead of one mega-block."""
    out = []
    cur = {"book": None, "ch": None, "v": None, "seg": 0}
    BLOCK = {"p", "q", "q1", "q2", "q3", "li", "li1", "li2", "pc", "pmo", "pm", "d", "m", "mi"}
    for event, elem in ET.iterparse(io.BytesIO(xml_bytes), events=("start", "end")):
        tag = local(elem.tag)
        if event == "start":
            if tag == "book":
                cur["book"] = (elem.get("id") or "").upper()
                cur["ch"] = cur["v"] = None; cur["seg"] = 0
            elif tag == "c":
                cur["ch"] = elem.get("id")
            elif tag == "v":
                cur["v"] = elem.get("id")
            elif tag in BLOCK:
                cur["seg"] += 1
        else:  # end
            if tag == "wj" and cur["book"] in BOOKS:
                text = " ".join("".join(elem.itertext()).split()).strip()
                if text:
                    out.append((cur["book"], cur["ch"], cur["v"], cur["seg"], text))
            if tag in ("book", "c", "p", "q", "wj", "v"):
                elem.clear()
    return out

def to_int(x):
    try: return int("".join(ch for ch in (x or "") if ch.isdigit()))
    except ValueError: return 0

def group_sayings(spans):
    """Merge consecutive verses bearing red letter into contiguous sayings.
    Spans may be 4-tuples (book, ch, v, text) or 5-tuples (book, ch, v, seg, text).
    When seg is present, sayings are additionally split at block boundaries.
    """
    # normalise to 5-tuples
    if spans and len(spans[0]) == 4:
        spans = [(b, c, v, 0, t) for b, c, v, t in spans]

    # collapse multiple wj in one verse+seg, keep first-seen order
    by_key = {}
    order = []
    for b, c, v, seg, t in spans:
        key = (b, to_int(c), to_int(v), seg)
        if key not in by_key:
            by_key[key] = []
            order.append(key)
        by_key[key].append(t)
    order.sort(key=lambda k: (BOOK_ORDER.index(k[0]) if k[0] in BOOK_ORDER else 99, k[1], k[2], k[3]))

    sayings, cur = [], None
    for key in order:
        b, c, v, seg = key
        text = " ".join(by_key[key])
        # merge only when same book, chapter, segment, and immediately next verse
        if cur and cur["b"] == b and cur["c"] == c and cur["seg"] == seg and v == cur["v_end"] + 1:
            cur["v_end"] = v
            cur["text"] += " " + text
        else:
            if cur: sayings.append(cur)
            cur = {"b": b, "c": c, "v_start": v, "v_end": v, "seg": seg, "text": text}
    if cur: sayings.append(cur)
    return sayings

def ref_str(s):
    name = BOOKS[s["b"]][0]
    if s["v_start"] == s["v_end"]:
        return f"{name} {s['c']}:{s['v_start']}"
    return f"{name} {s['c']}:{s['v_start']}-{s['v_end']}"

def write_xlsx(sayings, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    headers = ["ID","Scripture","Author of Book","Original (WEB)","Modern Rendering",
               "Situational Context","Sentiment","Audience Present","Approx. Age",
               "Location","Reason Present / Occasion","Reasoning Move"]
    wb = Workbook(); ws = wb.active; ws.title = "Sayings (full)"
    hf = PatternFill("solid", start_color="1F3864")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cfont = Font(name="Arial", size=10)
    thin = Side(style="thin", color="C9C9C9"); border = Border(thin,thin,thin,thin)
    wrap = Alignment(vertical="top", wrap_text=True)
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c); cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True); cell.border = border
    for i, s in enumerate(sayings, start=1):
        ws.append([i, ref_str(s), BOOKS[s["b"]][1], s["text"], "", "", "", "",
                   "~30-33 (ministry; verify)", "", "", ""])
    for r in range(2, ws.max_row+1):
        for c in range(1, len(headers)+1):
            cell = ws.cell(row=r, column=c); cell.font = cfont; cell.alignment = wrap; cell.border = border
            if r % 2 == 0: cell.fill = PatternFill("solid", start_color="EAF0F6")
    widths = [5,16,15,70,40,34,16,24,20,22,32,24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(path)

SELFTEST_USFX = b'''<usfx>
 <book id="MRK">
  <c id="1"/>
  <p><v id="14"/>Now after John was taken into custody, Jesus came into Galilee.
     <v id="15"/><wj>The time is fulfilled, and God's Kingdom is at hand! Repent, and believe in the Good News.</wj></p>
  <p><v id="16"/>Passing along by the sea, he saw Simon.
     <v id="17"/><wj>Come after me, and I will make you into fishers for men.</wj></p>
  <p><v id="40"/>A leper came to him.
     <v id="41"/>He stretched out his hand and said, <wj>I want to. Be made clean.</wj></p>
 </book>
</usfx>'''

def selftest():
    spans = parse_wj(SELFTEST_USFX)
    sayings = group_sayings(spans)
    refs = [ref_str(s) for s in sayings]
    # v16 is narrative (no <wj>), so it correctly breaks contiguity between 15 and 17
    assert refs == ["Mark 1:15", "Mark 1:17", "Mark 1:41"], refs
    assert sayings[0]["text"].startswith("The time is fulfilled")
    assert sayings[1]["text"] == "Come after me, and I will make you into fishers for men."
    assert sayings[2]["text"] == "I want to. Be made clean."
    print("SELF-TEST PASSED:", refs)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--usfx", help="path to a local USFX .xml (skips download)")
    ap.add_argument("--out", default="jesus_full_red_letter.xlsx")
    ap.add_argument("--selftest", action="store_true")
    args = ap.parse_args()
    if args.selftest:
        selftest(); return
    xml_bytes = get_usfx_bytes(args)
    spans = parse_wj(xml_bytes)
    sayings = group_sayings(spans)
    write_xlsx(sayings, args.out)
    print(f"Extracted {len(spans)} <wj> spans -> {len(sayings)} contiguous sayings")
    print(f"Wrote {args.out}")

if __name__ == "__main__":
    main()
