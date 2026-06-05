#!/usr/bin/env python3
"""
Convert the annotated red-letter corpus (jesus_full_red_letter.xlsx) into the
training/eval artifacts the digital-twin pipeline consumes:

  1. sft_style.jsonl       -- chat-format SFT records for the LoRA style adapter
  2. rag_corpus.jsonl      -- grounded passages for the retrieval index
  3. eval_heldout.jsonl    -- held-out sayings for style/grounding evaluation

WHY TWO OUTPUTS FROM ONE SHEET
------------------------------
The LoRA adapter learns *voice and reasoning form*; the RAG index supplies
*authoritative content*. Both are derived from the same annotated rows so they
stay consistent, but they are shaped very differently (see training_data_spec.md).

A row is "training-ready" only if BOTH 'Modern Rendering' and 'Reasoning Move'
are filled. Rows missing either are skipped and counted, so you can track
annotation progress against the 489-saying corpus.

USAGE
-----
    pip install openpyxl
    python build_training_jsonl.py --xlsx jesus_full_red_letter.xlsx --out-dir build/
    # held-out split fraction (default 0.10), deterministic by ID hash:
    python build_training_jsonl.py --xlsx ... --eval-frac 0.10
"""
from __future__ import annotations
import argparse, hashlib, json, os, re, sys

# Column order in the sheet's single header row.
COLS = [
    "ID", "Scripture", "Author of Book", "Original (WEB)", "Modern Rendering",
    "Situational Context", "Sentiment", "Audience Present", "Approx. Age",
    "Location", "Reason Present / Occasion", "Reasoning Move",
]

# Map a free-text "Reasoning Move" cell to a canonical move id (M01-M18) when
# possible. The annotation sheet allows prose; we normalize the leading token.
MOVE_RE = re.compile(r"\bM(\d{2})\b")

# The instruction that frames every SFT example. Kept identical at train and
# inference time so the adapter is conditioned on a stable system contract.
SYSTEM_PROMPT = (
    "You are a conversational mentor who responds as Jesus of Nazareth would, "
    "drawing only from his attested teachings and documented rhetorical methods. "
    "You speak directly and warmly in modern English, applying his characteristic "
    "reasoning moves to the questioner's situation. You never fabricate doctrine "
    "or invent sayings beyond the canonical record. When a question lies outside "
    "his attested words, you acknowledge it plainly and in his voice."
)


def clean(v) -> str:
    return ("" if v is None else str(v)).strip()


def is_ready(row: dict) -> bool:
    return bool(row["Modern Rendering"]) and bool(row["Reasoning Move"])


def move_id(cell: str) -> str:
    m = MOVE_RE.search(cell)
    return f"M{m.group(1)}" if m else ""


def in_eval(row_id: str, frac: float) -> bool:
    if frac <= 0:
        return False
    h = int(hashlib.sha1(row_id.encode()).hexdigest(), 16)
    return (h % 1000) < int(frac * 1000)


def build_sft_record(row: dict) -> dict:
    """One supervised example: the annotated context becomes the user turn,
    the modern rendering becomes the assistant turn. The reasoning-move tag is
    carried as metadata so it can be used for curriculum/weighting, not leaked
    into the visible prompt."""
    ref = row["Scripture"]
    ctx_bits = [b for b in (
        row["Situational Context"],
        f"Audience: {row['Audience Present']}" if row["Audience Present"] else "",
        f"Occasion: {row['Reason Present / Occasion']}" if row["Reason Present / Occasion"] else "",
    ) if b]
    context = " ".join(ctx_bits) if ctx_bits else "No additional context recorded."
    user = (
        f"Context: {context}\n"
        f"Render the following saying ({ref}) in present-day English, keeping its "
        f"original force and reasoning:\n“{row['Original (WEB)']}”"
    )
    return {
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user},
            {"role": "assistant", "content": row["Modern Rendering"]},
        ],
        "meta": {
            "id": row["ID"], "ref": ref,
            "move": move_id(row["Reasoning Move"]),
            "move_text": row["Reasoning Move"],
            "sentiment": row["Sentiment"], "audience": row["Audience Present"],
        },
    }


def build_rag_record(row: dict) -> dict:
    """One retrievable passage. The original WEB text is the ground truth;
    modern rendering and context are stored as searchable fields so a query can
    match either ancient or modern phrasing and still cite the canonical verse."""
    return {
        "id": f"wj-{row['ID']}",
        "ref": row["Scripture"],
        "book_author": row["Author of Book"],
        "text_original": row["Original (WEB)"],
        "text_modern": row["Modern Rendering"],
        "context": row["Situational Context"],
        "location": row["Location"],
        "occasion": row["Reason Present / Occasion"],
        "move": move_id(row["Reasoning Move"]),
        "translation": "World English Bible (public domain)",
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="jesus_full_red_letter.xlsx")
    ap.add_argument("--out-dir", default="build")
    ap.add_argument("--eval-frac", type=float, default=0.10)
    ap.add_argument("--sheet", default=None, help="sheet name; default = first")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(args.xlsx, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [clean(c) for c in next(rows_iter)]
    if header[:len(COLS)] != COLS:
        print(f"WARN: header mismatch.\n got: {header}\n want: {COLS}", file=sys.stderr)

    os.makedirs(args.out_dir, exist_ok=True)
    p_sft = os.path.join(args.out_dir, "sft_style.jsonl")
    p_rag = os.path.join(args.out_dir, "rag_corpus.jsonl")
    p_eval = os.path.join(args.out_dir, "eval_heldout.jsonl")

    n_total = n_ready = n_train = n_eval = n_rag = 0
    move_counts: dict[str, int] = {}

    with open(p_sft, "w") as f_sft, open(p_rag, "w") as f_rag, open(p_eval, "w") as f_eval:
        for raw in rows_iter:
            if raw is None or not any(raw):
                continue
            row = {COLS[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(COLS))}
            if not row["ID"]:
                continue
            n_total += 1
            # Every row with original text and a reference is retrievable.
            if row["Original (WEB)"]:
                f_rag.write(json.dumps(build_rag_record(row), ensure_ascii=False) + "\n")
                n_rag += 1
            if not is_ready(row):
                continue
            n_ready += 1
            mv = move_id(row["Reasoning Move"]) or "unmapped"
            move_counts[mv] = move_counts.get(mv, 0) + 1
            rec = build_sft_record(row)
            if in_eval(row["ID"], args.eval_frac):
                f_eval.write(json.dumps(rec, ensure_ascii=False) + "\n")
                n_eval += 1
            else:
                f_sft.write(json.dumps(rec, ensure_ascii=False) + "\n")
                n_train += 1

    print(f"rows seen ............ {n_total}")
    print(f"RAG passages ......... {n_rag}  -> {p_rag}")
    print(f"training-ready ....... {n_ready}  ({n_total - n_ready} awaiting annotation)")
    print(f"  SFT train .......... {n_train}  -> {p_sft}")
    print(f"  held-out eval ...... {n_eval}  -> {p_eval}")
    print("move distribution (training-ready):")
    for mv in sorted(move_counts):
        print(f"  {mv:>8}: {move_counts[mv]}")
    if n_ready < 200:
        print("\nNOTE: <200 training-ready rows. Style LoRA is viable but thin; "
              "prioritize finishing annotation and/or augment per training_data_spec.md.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
